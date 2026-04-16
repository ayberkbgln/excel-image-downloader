"""
Excel Image Downloader v3.0
- TR / EN dil secici
- About penceresi (iletisim linkleri)
- Ayar kaydetme (config.json)
- Cikti klasorunu ac butonu
- Drag & Drop (opsiyonel - tkinterdnd2 varsa)
- Daha guzel UI
"""
import os, re, sys, json, threading, time, queue, webbrowser, subprocess
from pathlib import Path
from urllib.parse import urlparse
from urllib.request import urlopen, Request
from concurrent.futures import ThreadPoolExecutor, as_completed

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook

# Opsiyonel drag&drop
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    DND_AVAILABLE = True
except ImportError:
    DND_AVAILABLE = False


APP_VER = "v3.0"
MAX_LOG_LINES = 500
CONFIG_DIR = Path.home() / ".excel-image-downloader"
CONFIG_FILE = CONFIG_DIR / "config.json"

# --------------------- I18N ---------------------
TRANSLATIONS = {
    "tr": {
        "title": "Resim İndirici",
        "subtitle": "Excel'deki URL linklerinden toplu resim indirici",
        "excel_file": "EXCEL DOSYASI",
        "output_folder": "ÇIKTI KLASÖRÜ",
        "parallel": "PARALEL İNDİRME",
        "thread_unit": "thread",
        "skip_existing": "  Var olan dosyaları atla (resume)",
        "single_folder": "  Tek klasöre indir (alt klasör yok)",
        "browse": "📁 Gözat",
        "start": "▶  İNDİRMEYİ BAŞLAT",
        "stop": "⏹  DURDUR",
        "downloading": "⏳  İNDİRİLİYOR...",
        "pause": "Durdur",
        "ready": "● Hazır",
        "downloading_status": "● İndiriliyor...",
        "stopping": "● Durduruluyor...",
        "done": "● Bitti",
        "live_log": "Canlı Log (son 500 satır)",
        "clear": "Temizle",
        "open_folder": "📂 Klasörü Aç",
        "about": "ℹ Hakkında",
        "err_file_invalid": "Geçerli bir Excel dosyası seç.",
        "err_no_output": "Çıktı klasörü seç.",
        "pick_excel": "Excel Seç",
        "pick_output": "Çıktı Klasörü",
        "drop_hint": "  📎 Excel dosyasını buraya sürükle-bırak",
        "no_urls": "URL bulunamadı.",
        "reading_excel": "► Excel okunuyor...",
        "found_urls": "► {n} URL bulundu. {t} paralel thread • {m} modunda başlıyor...",
        "mode_single": "tek klasör",
        "mode_grouped": "kod bazlı klasör",
        "ok": "OK",
        "skipped": "ATLANAN",
        "failed": "HATA",
        "total": "Toplam",
        "existing": "mevcut",
        "stopping_log": "⚠ Durduruluyor... (aktif indirmeler bittikten sonra)",
        "errors_written": "⚠ {n} hata 'hatalar.txt' dosyasına yazıldı",
        "completed": "Tamamlandı",
        "completed_msg": "İndirme bitti.\n\nToplam: {total}\nBaşarılı: {ok}\nAtlanan (mevcut): {skipped}\nHata: {fail}",
        "errors_location": "\n\nHatalar: {path}",
        "remaining": "kalan",
        "error": "Hata",
        "about_title": "Hakkında",
        "about_dev": "Geliştirici",
        "about_contact": "İletişim",
        "about_desc": "Excel dosyalarındaki URL linklerinden toplu resim indirici.\nModern dark UI, paralel indirme, resume desteği.",
        "about_close": "Kapat",
        "lang_tooltip": "Dil değiştir",
    },
    "en": {
        "title": "Image Downloader",
        "subtitle": "Bulk image downloader from Excel URL links",
        "excel_file": "EXCEL FILE",
        "output_folder": "OUTPUT FOLDER",
        "parallel": "PARALLEL DOWNLOADS",
        "thread_unit": "threads",
        "skip_existing": "  Skip existing files (resume)",
        "single_folder": "  Single folder (no subfolders)",
        "browse": "📁 Browse",
        "start": "▶  START DOWNLOAD",
        "stop": "⏹  STOP",
        "downloading": "⏳  DOWNLOADING...",
        "pause": "Stop",
        "ready": "● Ready",
        "downloading_status": "● Downloading...",
        "stopping": "● Stopping...",
        "done": "● Done",
        "live_log": "Live Log (last 500 lines)",
        "clear": "Clear",
        "open_folder": "📂 Open Folder",
        "about": "ℹ About",
        "err_file_invalid": "Select a valid Excel file.",
        "err_no_output": "Select an output folder.",
        "pick_excel": "Select Excel",
        "pick_output": "Output Folder",
        "drop_hint": "  📎 Drag & drop Excel file here",
        "no_urls": "No URL found.",
        "reading_excel": "► Reading Excel...",
        "found_urls": "► {n} URLs found. Starting with {t} threads in {m} mode...",
        "mode_single": "single folder",
        "mode_grouped": "grouped folders",
        "ok": "OK",
        "skipped": "SKIPPED",
        "failed": "FAILED",
        "total": "Total",
        "existing": "existing",
        "stopping_log": "⚠ Stopping... (after active downloads finish)",
        "errors_written": "⚠ {n} errors written to 'hatalar.txt'",
        "completed": "Completed",
        "completed_msg": "Download finished.\n\nTotal: {total}\nSuccess: {ok}\nSkipped (existing): {skipped}\nFailed: {fail}",
        "errors_location": "\n\nErrors: {path}",
        "remaining": "remaining",
        "error": "Error",
        "about_title": "About",
        "about_dev": "Developer",
        "about_contact": "Contact",
        "about_desc": "Bulk image downloader from Excel URL links.\nModern dark UI, parallel downloads, resume support.",
        "about_close": "Close",
        "lang_tooltip": "Change language",
    },
}

# --------------------- RENKLER ---------------------
BG        = "#0f172a"
BG_CARD   = "#1e293b"
BG_INPUT  = "#0b1220"
BG_HOVER  = "#334155"
BORDER    = "#334155"
FG        = "#e2e8f0"
FG_MUTED  = "#94a3b8"
ACCENT    = "#3b82f6"
ACCENT_HO = "#2563eb"
OK_GREEN  = "#10b981"
ERR_RED   = "#ef4444"
WARN_YEL  = "#f59e0b"

# --------------------- UTILS ---------------------
def resource_path(rel):
    try:
        base = sys._MEIPASS
    except Exception:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, rel)


def download(url, dest, timeout=30):
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=timeout) as r, open(dest, "wb") as f:
        f.write(r.read())


def safe_name(s):
    return re.sub(r'[<>:"/\\|?*]', "_", str(s)).strip()


def load_config():
    try:
        return json.loads(CONFIG_FILE.read_text("utf-8"))
    except Exception:
        return {}


def save_config(cfg):
    try:
        CONFIG_DIR.mkdir(exist_ok=True)
        CONFIG_FILE.write_text(json.dumps(cfg, indent=2), "utf-8")
    except Exception:
        pass


def open_folder(path):
    p = str(Path(path).resolve())
    if sys.platform == "win32":
        os.startfile(p)
    elif sys.platform == "darwin":
        subprocess.Popen(["open", p])
    else:
        subprocess.Popen(["xdg-open", p])


# --------------------- WIDGETS ---------------------
class HoverButton(tk.Canvas):
    def __init__(self, parent, text, command, bg=ACCENT, hover=ACCENT_HO,
                 fg="white", width=200, height=44, radius=10,
                 font=("Segoe UI", 10, "bold")):
        super().__init__(parent, width=width, height=height,
                         bg=parent["bg"], bd=0, highlightthickness=0)
        self.command = command
        self.bg_c = bg; self.hover_c = hover; self.fg = fg
        self.radius = radius; self.w = width; self.h = height
        self.text = text; self.font = font
        self._enabled = True
        self._draw(bg)
        self.bind("<Enter>", lambda e: self._enabled and self._draw(self.hover_c))
        self.bind("<Leave>", lambda e: self._enabled and self._draw(self.bg_c))
        self.bind("<Button-1>", lambda e: self._enabled and self.command and self.command())

    def _round_rect(self, x1, y1, x2, y2, r, **kw):
        pts = [x1+r,y1, x2-r,y1, x2,y1, x2,y1+r, x2,y2-r,
               x2,y2, x2-r,y2, x1+r,y2, x1,y2, x1,y2-r,
               x1,y1+r, x1,y1]
        return self.create_polygon(pts, smooth=True, **kw)

    def _draw(self, color):
        self.delete("all")
        self._round_rect(1, 1, self.w-1, self.h-1, self.radius,
                         fill=color, outline=color)
        self.create_text(self.w//2, self.h//2, text=self.text,
                         fill=self.fg, font=self.font)

    def set_text(self, t):
        self.text = t; self._draw(self.bg_c)

    def set_enabled(self, val):
        self._enabled = val
        self._draw(self.bg_c if val else "#475569")

    def configure_colors(self, bg=None, hover=None):
        if bg: self.bg_c = bg
        if hover: self.hover_c = hover
        self._draw(self.bg_c if self._enabled else "#475569")


class LangToggle(tk.Canvas):
    """Yuvarlatilmis dil secici buton (TR / EN)."""
    def __init__(self, parent, current, on_change):
        w, h = 96, 30
        super().__init__(parent, width=w, height=h,
                         bg=parent["bg"], bd=0, highlightthickness=0)
        self.w = w; self.h = h
        self.on_change = on_change
        self.current = current
        self._draw()
        self.bind("<Button-1>", self._click)

    def _round_rect(self, x1, y1, x2, y2, r, **kw):
        pts = [x1+r,y1, x2-r,y1, x2,y1, x2,y1+r, x2,y2-r,
               x2,y2, x2-r,y2, x1+r,y2, x1,y2, x1,y2-r,
               x1,y1+r, x1,y1]
        return self.create_polygon(pts, smooth=True, **kw)

    def _draw(self):
        self.delete("all")
        # Arka plan
        self._round_rect(1, 1, self.w-1, self.h-1, self.h//2,
                         fill=BG_CARD, outline=BORDER)
        # Aktif kaydirma topu
        idx = 0 if self.current == "tr" else 1
        half = self.w // 2
        x1 = 2 + idx * (half - 2)
        self._round_rect(x1, 3, x1 + half, self.h - 3, self.h//2 - 3,
                         fill=ACCENT, outline=ACCENT)
        # Yazilar
        self.create_text(self.w // 4, self.h // 2, text="🇹🇷 TR",
                         fill="white" if self.current == "tr" else FG_MUTED,
                         font=("Segoe UI", 9, "bold"))
        self.create_text(3 * self.w // 4, self.h // 2, text="🇬🇧 EN",
                         fill="white" if self.current == "en" else FG_MUTED,
                         font=("Segoe UI", 9, "bold"))

    def _click(self, e):
        new = "en" if e.x > self.w // 2 else "tr"
        if new != self.current:
            self.current = new
            self._draw()
            self.on_change(new)

    def set_lang(self, lang):
        self.current = lang
        self._draw()


# --------------------- APP ---------------------
class App:
    def __init__(self, root):
        self.root = root
        self.cfg = load_config()
        self.lang = self.cfg.get("lang", "tr")

        root.title(f"Image Downloader {APP_VER}")
        root.geometry("860x740")
        root.minsize(800, 660)
        root.configure(bg=BG)
        try:
            root.iconbitmap(resource_path("icon.ico"))
        except Exception:
            pass

        # State
        self.xlsx_path = tk.StringVar(value=self.cfg.get("last_xlsx", ""))
        self.out_path  = tk.StringVar(value=self.cfg.get("last_out", ""))
        self.threads_v = tk.IntVar(value=self.cfg.get("threads", 10))
        self.skip_existing = tk.BooleanVar(value=self.cfg.get("skip", True))
        self.single_folder = tk.BooleanVar(value=self.cfg.get("single", False))
        self.running = False
        self.stop_flag = threading.Event()
        self.log_queue = queue.Queue()
        self.log_count = 0
        self.last_output = None

        self._setup_style()
        self._build_ui()
        self._apply_lang()
        self._poll_log()
        self._setup_dnd()

        # Uygulamayi kapatirken ayarlari kaydet
        root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ---- I18N ----
    def t(self, key, **kw):
        txt = TRANSLATIONS[self.lang].get(key, key)
        return txt.format(**kw) if kw else txt

    def _apply_lang(self):
        self.title_lbl.config(text="📥  " + self.t("title"))
        self.subtitle_lbl.config(text=f"{self.t('subtitle')}  •  {APP_VER}")
        self.excel_lbl.config(text=self.t("excel_file"))
        self.out_lbl.config(text=self.t("output_folder"))
        self.par_lbl.config(text=self.t("parallel"))
        self.thread_val_lbl.config(
            text=f"{int(self.threads_v.get())} {self.t('thread_unit')}")
        self.skip_cb.config(text=self.t("skip_existing"))
        self.single_cb.config(text=self.t("single_folder"))
        self.browse_xlsx.set_text(self.t("browse"))
        self.browse_out.set_text(self.t("browse"))
        self.start_btn.set_text(
            self.t("downloading") if self.running else self.t("start"))
        self.stop_btn.set_text(self.t("stop"))
        self.log_hdr.config(text=self.t("live_log"))
        self.clear_btn.config(text=self.t("clear"))
        if self.last_output:
            self.open_folder_btn.set_text(self.t("open_folder"))
        self.about_btn.set_text(self.t("about"))
        if not self.running:
            self.status.config(text=self.t("ready"))
        self.root.title(f"{self.t('title')} {APP_VER}")

    def _on_lang_change(self, lang):
        self.lang = lang
        self.cfg["lang"] = lang
        save_config(self.cfg)
        self._apply_lang()

    # ---- UI ----
    def _setup_style(self):
        style = ttk.Style()
        try: style.theme_use("clam")
        except: pass
        style.configure("Modern.Horizontal.TProgressbar",
            troughcolor=BG_INPUT, background=ACCENT,
            bordercolor=BG_INPUT, lightcolor=ACCENT, darkcolor=ACCENT,
            thickness=14)

    def _card(self, parent):
        return tk.Frame(parent, bg=BG_CARD, highlightbackground=BORDER,
                        highlightthickness=1)

    def _entry(self, parent, var):
        return tk.Entry(parent, textvariable=var, bg=BG_INPUT, fg=FG,
            insertbackground=FG, relief="flat", font=("Segoe UI", 10),
            highlightthickness=1, highlightbackground=BORDER,
            highlightcolor=ACCENT)

    def _build_ui(self):
        # === HEADER ===
        h = tk.Frame(self.root, bg=BG)
        h.pack(fill="x", padx=24, pady=(18, 4))

        left = tk.Frame(h, bg=BG)
        left.pack(side="left", fill="x", expand=True)
        self.title_lbl = tk.Label(left, text="", bg=BG, fg=FG,
                                  font=("Segoe UI", 20, "bold"))
        self.title_lbl.pack(anchor="w")
        self.subtitle_lbl = tk.Label(left, text="", bg=BG, fg=FG_MUTED,
                                     font=("Segoe UI", 10))
        self.subtitle_lbl.pack(anchor="w")

        right = tk.Frame(h, bg=BG)
        right.pack(side="right")
        self.lang_toggle = LangToggle(right, self.lang, self._on_lang_change)
        self.lang_toggle.pack(pady=4)

        # === AYARLAR ===
        card = self._card(self.root)
        card.pack(fill="x", padx=24, pady=10)
        inner = tk.Frame(card, bg=BG_CARD)
        inner.pack(fill="x", padx=18, pady=16)

        self.excel_lbl = tk.Label(inner, text="", bg=BG_CARD, fg=FG_MUTED,
            font=("Segoe UI", 8, "bold"))
        self.excel_lbl.pack(anchor="w")
        r1 = tk.Frame(inner, bg=BG_CARD)
        r1.pack(fill="x", pady=(4, 12))
        self.xlsx_entry = self._entry(r1, self.xlsx_path)
        self.xlsx_entry.pack(side="left", fill="x", expand=True, ipady=6)
        self.browse_xlsx = HoverButton(r1, "", self.pick_xlsx,
            bg=BG_HOVER, hover="#475569", width=110, height=34)
        self.browse_xlsx.pack(side="left", padx=(8, 0))

        self.out_lbl = tk.Label(inner, text="", bg=BG_CARD, fg=FG_MUTED,
            font=("Segoe UI", 8, "bold"))
        self.out_lbl.pack(anchor="w")
        r2 = tk.Frame(inner, bg=BG_CARD)
        r2.pack(fill="x", pady=(4, 12))
        self.out_entry = self._entry(r2, self.out_path)
        self.out_entry.pack(side="left", fill="x", expand=True, ipady=6)
        self.browse_out = HoverButton(r2, "", self.pick_out,
            bg=BG_HOVER, hover="#475569", width=110, height=34)
        self.browse_out.pack(side="left", padx=(8, 0))

        # Opsiyonlar
        opt = tk.Frame(inner, bg=BG_CARD)
        opt.pack(fill="x", pady=(4, 0))

        tl = tk.Frame(opt, bg=BG_CARD)
        tl.pack(side="left")
        self.par_lbl = tk.Label(tl, text="", bg=BG_CARD, fg=FG_MUTED,
            font=("Segoe UI", 8, "bold"))
        self.par_lbl.pack(anchor="w")
        sl_wrap = tk.Frame(tl, bg=BG_CARD)
        sl_wrap.pack(anchor="w", pady=(4, 0))
        self.thread_val_lbl = tk.Label(sl_wrap, text="", bg=BG_CARD, fg=FG,
            font=("Segoe UI", 9, "bold"), width=12, anchor="w")
        self.thread_val_lbl.pack(side="right", padx=(8, 0))
        sc = ttk.Scale(sl_wrap, from_=1, to=32, orient="horizontal",
            variable=self.threads_v, length=240,
            command=lambda v: self.thread_val_lbl.config(
                text=f"{int(float(v))} {self.t('thread_unit')}"))
        sc.pack(side="left")

        sk = tk.Frame(opt, bg=BG_CARD)
        sk.pack(side="right")
        cb_kw = dict(bg=BG_CARD, fg=FG, selectcolor=BG_INPUT,
            activebackground=BG_CARD, activeforeground=FG,
            font=("Segoe UI", 9), bd=0, highlightthickness=0, anchor="w")
        self.skip_cb = tk.Checkbutton(sk, text="",
            variable=self.skip_existing, **cb_kw)
        self.skip_cb.pack(anchor="w", pady=(14, 0))
        self.single_cb = tk.Checkbutton(sk, text="",
            variable=self.single_folder, **cb_kw)
        self.single_cb.pack(anchor="w")

        # === BUTTONS ===
        bf = tk.Frame(self.root, bg=BG)
        bf.pack(fill="x", padx=24, pady=(14, 6))
        self.start_btn = HoverButton(bf, "", self.start,
            bg=ACCENT, hover=ACCENT_HO, width=640, height=52,
            radius=12, font=("Segoe UI", 12, "bold"))
        self.start_btn.pack(side="left", fill="x", expand=True)
        self.stop_btn = HoverButton(bf, "", self.stop,
            bg=ERR_RED, hover="#dc2626", width=140, height=52,
            radius=12, font=("Segoe UI", 11, "bold"))
        self.stop_btn.pack(side="left", padx=(8, 0))
        self.stop_btn.set_enabled(False)

        # === PROGRESS ===
        pf = tk.Frame(self.root, bg=BG)
        pf.pack(fill="x", padx=24, pady=(8, 4))
        self.progress = ttk.Progressbar(pf,
            style="Modern.Horizontal.TProgressbar", mode="determinate")
        self.progress.pack(fill="x")

        st = tk.Frame(self.root, bg=BG)
        st.pack(fill="x", padx=24, pady=(4, 6))
        self.status = tk.Label(st, text="", bg=BG, fg=FG_MUTED,
            font=("Segoe UI", 9), anchor="w")
        self.status.pack(side="left")
        self.counter = tk.Label(st, text="", bg=BG, fg=FG,
            font=("Segoe UI", 9, "bold"), anchor="e")
        self.counter.pack(side="right")

        # === LOG ===
        lc = self._card(self.root)
        lc.pack(fill="both", expand=True, padx=24, pady=(6, 6))
        lh = tk.Frame(lc, bg=BG_CARD)
        lh.pack(fill="x", padx=14, pady=(10, 4))
        self.log_hdr = tk.Label(lh, text="", bg=BG_CARD, fg=FG_MUTED,
            font=("Segoe UI", 8, "bold"))
        self.log_hdr.pack(side="left")

        self.open_folder_btn = HoverButton(lh, "📂", self._open_output,
            bg=BG_CARD, hover=BG_HOVER, fg=FG_MUTED, width=120, height=22,
            radius=6, font=("Segoe UI", 8, "bold"))
        # Baslangicta gizli; tamamlaninca goster

        self.clear_btn = tk.Button(lh, text="", bg=BG_CARD, fg=FG_MUTED,
            bd=0, activebackground=BG_CARD, activeforeground=FG,
            font=("Segoe UI", 8), cursor="hand2",
            command=lambda: self.log.delete("1.0", "end"))
        self.clear_btn.pack(side="right")

        lw = tk.Frame(lc, bg=BG_INPUT)
        lw.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        self.log = tk.Text(lw, bg=BG_INPUT, fg="#cbd5e1", bd=0,
            insertbackground=FG, font=("Consolas", 9),
            relief="flat", padx=10, pady=8, wrap="none")
        self.log.pack(side="left", fill="both", expand=True)
        sb = tk.Scrollbar(lw, command=self.log.yview, bg=BG_CARD)
        sb.pack(side="right", fill="y")
        self.log.config(yscrollcommand=sb.set)
        self.log.tag_config("ok",   foreground=OK_GREEN)
        self.log.tag_config("err",  foreground=ERR_RED)
        self.log.tag_config("info", foreground=ACCENT)
        self.log.tag_config("warn", foreground=WARN_YEL)

        # === FOOTER ===
        ff = tk.Frame(self.root, bg=BG)
        ff.pack(side="bottom", fill="x", padx=24, pady=(0, 12))
        tk.Label(ff, text="Ayberk Bağlan", bg=BG, fg=FG_MUTED,
                 font=("Segoe UI", 9, "italic")).pack(side="left")
        self.about_btn = HoverButton(ff, "", self.show_about,
            bg=BG, hover=BG_CARD, fg=FG_MUTED, width=110, height=26,
            radius=6, font=("Segoe UI", 9))
        self.about_btn.pack(side="right")

    # ---- Drag & Drop ----
    def _setup_dnd(self):
        if not DND_AVAILABLE:
            return
        try:
            self.xlsx_entry.drop_target_register(DND_FILES)
            self.xlsx_entry.dnd_bind("<<Drop>>", self._on_drop)
        except Exception:
            pass

    def _on_drop(self, e):
        data = e.data.strip("{}").strip('"')
        if data.lower().endswith((".xlsx", ".xlsm")):
            self.xlsx_path.set(data)
            if not self.out_path.get():
                self.out_path.set(str(Path(data).parent / "indirilen"))

    # ---- File pick ----
    def pick_xlsx(self):
        p = filedialog.askopenfilename(title=self.t("pick_excel"),
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("All", "*.*")])
        if p:
            self.xlsx_path.set(p)
            if not self.out_path.get():
                self.out_path.set(str(Path(p).parent / "indirilen"))

    def pick_out(self):
        p = filedialog.askdirectory(title=self.t("pick_output"))
        if p:
            self.out_path.set(p)

    def _open_output(self):
        if self.last_output and Path(self.last_output).exists():
            open_folder(self.last_output)

    # ---- ABOUT DIALOG ----
    def show_about(self):
        AboutDialog(self.root, self)

    # ---- Log ----
    def log_write(self, msg, tag=None):
        self.log_queue.put((msg, tag))

    def _poll_log(self):
        try:
            n = 0
            while n < 200:
                msg, tag = self.log_queue.get_nowait()
                self.log.insert("end", msg + "\n", tag or ())
                self.log_count += 1; n += 1
        except queue.Empty:
            pass
        if self.log_count > MAX_LOG_LINES:
            ex = self.log_count - MAX_LOG_LINES
            self.log.delete("1.0", f"{ex + 1}.0")
            self.log_count = MAX_LOG_LINES
        self.log.see("end")
        self.root.after(80, self._poll_log)

    def set_status(self, msg, color=FG_MUTED):
        self.status.config(text=msg, fg=color)

    # ---- START/STOP ----
    def start(self):
        if self.running:
            return
        xlsx = self.xlsx_path.get().strip()
        out  = self.out_path.get().strip()
        if not xlsx or not Path(xlsx).exists():
            messagebox.showerror(self.t("error"), self.t("err_file_invalid"))
            return
        if not out:
            messagebox.showerror(self.t("error"), self.t("err_no_output"))
            return
        # Ayarlari kaydet
        self.cfg.update({
            "last_xlsx": xlsx, "last_out": out,
            "threads": int(self.threads_v.get()),
            "skip": bool(self.skip_existing.get()),
            "single": bool(self.single_folder.get()),
            "lang": self.lang,
        })
        save_config(self.cfg)

        self.running = True
        self.stop_flag.clear()
        self.start_btn.set_enabled(False)
        self.start_btn.set_text(self.t("downloading"))
        self.stop_btn.set_enabled(True)
        self.open_folder_btn.pack_forget()
        self.log.delete("1.0", "end")
        self.log_count = 0
        self.last_output = out
        threading.Thread(target=self._worker, args=(xlsx, out),
                         daemon=True).start()

    def stop(self):
        if not self.running:
            return
        self.stop_flag.set()
        self.log_write(self.t("stopping_log"), "warn")
        self.set_status(self.t("stopping"), WARN_YEL)

    # ---- WORKER ----
    def _one(self, kod, url, out_p, skip, single):
        if self.stop_flag.is_set():
            return ("skip", kod, url, "stopped")
        m = re.search(r"Product/(.+)$", url, re.IGNORECASE)
        name = m.group(1) if m else os.path.basename(urlparse(url).path)
        name = name.replace("/", "_")
        if single:
            dest = out_p / name
        else:
            klasor = out_p / kod
            klasor.mkdir(exist_ok=True)
            dest = klasor / name
        if skip and dest.exists() and dest.stat().st_size > 0:
            return ("skip", kod, name, "existing")
        try:
            download(url, dest)
            return ("ok", kod, name, "")
        except Exception as e:
            return ("err", kod, url, str(e))

    def _worker(self, xlsx, out):
        try:
            out_p = Path(out)
            out_p.mkdir(parents=True, exist_ok=True)
            self.log_write(self.t("reading_excel"), "info")
            wb = load_workbook(xlsx, data_only=True, read_only=True)
            ws = wb.active

            jobs = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                kod = safe_name(row[0])
                for cell in row[1:]:
                    if not cell: continue
                    url = str(cell).strip()
                    if url.lower().startswith("http"):
                        jobs.append((kod, url))
            wb.close()

            total = len(jobs)
            if total == 0:
                self.log_write(self.t("no_urls"), "err")
                self._done(0, 0, 0, 0, None)
                return

            threads = max(1, int(self.threads_v.get()))
            skip = bool(self.skip_existing.get())
            single = bool(self.single_folder.get())
            mode = self.t("mode_single") if single else self.t("mode_grouped")
            self.log_write(
                self.t("found_urls", n=total, t=threads, m=mode) + "\n",
                "info")
            self.progress.config(maximum=total, value=0)

            ok = fail = skipped = done = 0
            errors = []
            start_ts = time.time()

            with ThreadPoolExecutor(max_workers=threads) as ex:
                futs = [ex.submit(self._one, k, u, out_p, skip, single)
                        for k, u in jobs]
                for fut in as_completed(futs):
                    status, kod, name_or_url, extra = fut.result()
                    done += 1
                    if status == "ok":
                        ok += 1
                        if ok % 5 == 0 or done < 20:
                            self.log_write(f"✓ {kod}/{name_or_url}", "ok")
                    elif status == "skip":
                        skipped += 1
                        if extra == "stopped": pass
                        elif skipped <= 10 or skipped % 50 == 0:
                            self.log_write(
                                f"↷ {kod}/{name_or_url} ({self.t('existing')})",
                                "warn")
                    else:
                        fail += 1
                        errors.append(f"{name_or_url}\t{extra}")
                        self.log_write(f"✗ {name_or_url}  ({extra})", "err")

                    self.progress.config(value=done)
                    elapsed = time.time() - start_ts
                    rate = done / elapsed if elapsed > 0 else 0
                    eta = (total - done) / rate if rate > 0 else 0
                    self.set_status(
                        f"{self.t('downloading_status')}  {done}/{total}  •  "
                        f"{rate:.1f}/sn  •  {self.t('remaining')} ~{self._fmt_time(eta)}",
                        ACCENT)
                    self.counter.config(
                        text=f"{self.t('ok')} {ok}  •  {self.t('skipped')} {skipped}  •  {self.t('failed')} {fail}")

            err_file = None
            if errors:
                err_file = out_p / "hatalar.txt"
                err_file.write_text("\n".join(errors), encoding="utf-8")
                self.log_write(self.t("errors_written", n=len(errors)),
                               "warn")
            self._done(total, ok, fail, skipped, err_file)

        except Exception as e:
            self.log_write(f"[CRITICAL] {e}", "err")
            messagebox.showerror(self.t("error"), str(e))
            self._reset_btn()

    def _fmt_time(self, s):
        s = int(s)
        if s < 60: return f"{s}s"
        if s < 3600: return f"{s//60}m {s%60}s"
        return f"{s//3600}h {(s%3600)//60}m"

    def _done(self, total, ok, fail, skipped, err_file):
        color = OK_GREEN if fail == 0 else (WARN_YEL if ok > 0 else ERR_RED)
        self.set_status(self.t("done"), color)
        self.counter.config(
            text=f"{self.t('total')} {total}  •  "
                 f"{self.t('ok')} {ok}  •  "
                 f"{self.t('skipped')} {skipped}  •  "
                 f"{self.t('failed')} {fail}")
        self.log_write(
            f"\n═══ {self.t('done').replace('●','').strip()} ═══  "
            f"{self.t('total')}={total}  {self.t('ok')}={ok}  "
            f"{self.t('skipped')}={skipped}  {self.t('failed')}={fail}",
            "info")
        msg = self.t("completed_msg",
                     total=total, ok=ok, skipped=skipped, fail=fail)
        if err_file:
            msg += self.t("errors_location", path=str(err_file))
        # Klasoru ac butonunu goster
        self.open_folder_btn.set_text(self.t("open_folder"))
        self.open_folder_btn.pack(side="right", padx=(0, 8))
        messagebox.showinfo(self.t("completed"), msg)
        self._reset_btn()

    def _reset_btn(self):
        self.running = False
        self.stop_flag.clear()
        self.start_btn.set_enabled(True)
        self.start_btn.set_text(self.t("start"))
        self.stop_btn.set_enabled(False)

    def _on_close(self):
        self.cfg.update({
            "lang": self.lang,
            "threads": int(self.threads_v.get()),
            "skip": bool(self.skip_existing.get()),
            "single": bool(self.single_folder.get()),
            "last_xlsx": self.xlsx_path.get(),
            "last_out": self.out_path.get(),
        })
        save_config(self.cfg)
        self.root.destroy()


# --------------------- ABOUT DIALOG ---------------------
class AboutDialog:
    LINKS = [
        ("🐦", "X / Twitter", "@yulewiz",
         "https://x.com/yulewiz"),
        ("💼", "LinkedIn", "in/ayberkbaglan",
         "https://www.linkedin.com/in/ayberkbaglan/"),
        ("🐙", "GitHub", "@ayberkbgln",
         "https://github.com/ayberkbgln"),
        ("✉", "Email", "ayberkbaglan@gmail.com",
         "mailto:ayberkbaglan@gmail.com"),
    ]

    def __init__(self, parent, app):
        self.app = app
        self.top = tk.Toplevel(parent)
        self.top.title(app.t("about_title"))
        self.top.geometry("420x480")
        self.top.resizable(False, False)
        self.top.configure(bg=BG)
        self.top.transient(parent)
        try:
            self.top.iconbitmap(resource_path("icon.ico"))
        except Exception:
            pass

        # Ortala
        parent.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() - 420) // 2
        y = parent.winfo_y() + (parent.winfo_height() - 480) // 2
        self.top.geometry(f"+{x}+{y}")

        self._build()
        self.top.grab_set()

    def _build(self):
        # Ust - icon
        try:
            from PIL import Image, ImageTk
            img = Image.open(resource_path("icon.ico"))
            img.thumbnail((80, 80))
            self.icon_img = ImageTk.PhotoImage(img)
            tk.Label(self.top, image=self.icon_img, bg=BG).pack(pady=(22, 8))
        except Exception:
            tk.Label(self.top, text="📥", bg=BG, fg=FG,
                     font=("Segoe UI", 40)).pack(pady=(22, 8))

        tk.Label(self.top, text=self.app.t("title"),
                 bg=BG, fg=FG, font=("Segoe UI", 16, "bold")).pack()
        tk.Label(self.top, text=APP_VER, bg=BG, fg=FG_MUTED,
                 font=("Segoe UI", 9)).pack()
        tk.Label(self.top, text=self.app.t("about_desc"),
                 bg=BG, fg=FG_MUTED, font=("Segoe UI", 9),
                 justify="center").pack(pady=(8, 12))

        # Ayraç
        tk.Frame(self.top, bg=BORDER, height=1).pack(
            fill="x", padx=30, pady=(4, 12))

        # Gelistirici
        tk.Label(self.top,
                 text=f"{self.app.t('about_dev').upper()}",
                 bg=BG, fg=FG_MUTED,
                 font=("Segoe UI", 8, "bold")).pack()
        tk.Label(self.top, text="Ayberk Bağlan", bg=BG, fg=FG,
                 font=("Segoe UI", 13, "bold")).pack(pady=(2, 10))

        # Linkler
        tk.Label(self.top, text=self.app.t("about_contact").upper(),
                 bg=BG, fg=FG_MUTED,
                 font=("Segoe UI", 8, "bold")).pack(pady=(4, 6))

        lf = tk.Frame(self.top, bg=BG)
        lf.pack(fill="x", padx=30)
        for emoji, name, handle, url in self.LINKS:
            row = tk.Frame(lf, bg=BG_CARD, highlightbackground=BORDER,
                           highlightthickness=1, cursor="hand2")
            row.pack(fill="x", pady=3)
            inner = tk.Frame(row, bg=BG_CARD, cursor="hand2")
            inner.pack(fill="x", padx=10, pady=6)
            tk.Label(inner, text=emoji, bg=BG_CARD, fg=FG,
                     font=("Segoe UI", 14), cursor="hand2").pack(side="left")
            txt_frame = tk.Frame(inner, bg=BG_CARD, cursor="hand2")
            txt_frame.pack(side="left", padx=(10, 0), fill="x", expand=True)
            tk.Label(txt_frame, text=name, bg=BG_CARD, fg=FG,
                     font=("Segoe UI", 9, "bold"),
                     cursor="hand2").pack(anchor="w")
            tk.Label(txt_frame, text=handle, bg=BG_CARD, fg=FG_MUTED,
                     font=("Segoe UI", 8),
                     cursor="hand2").pack(anchor="w")
            tk.Label(inner, text="↗", bg=BG_CARD, fg=FG_MUTED,
                     font=("Segoe UI", 11), cursor="hand2").pack(side="right")

            # Hover & click
            def bind_all(widget, u=url, r=row):
                widget.bind("<Button-1>", lambda e: webbrowser.open(u))
                widget.bind("<Enter>", lambda e: self._hover(r, True))
                widget.bind("<Leave>", lambda e: self._hover(r, False))
            for w in [row, inner, txt_frame] + list(inner.winfo_children()) + list(txt_frame.winfo_children()):
                bind_all(w)

        # Kapat
        HoverButton(self.top, self.app.t("about_close"),
                    self.top.destroy, bg=BG_HOVER, hover="#475569",
                    width=120, height=34, radius=8).pack(pady=16)

    def _hover(self, widget, enter):
        color = BG_HOVER if enter else BG_CARD
        widget.config(bg=color)
        for w in widget.winfo_children():
            try:
                w.config(bg=color)
                for w2 in w.winfo_children():
                    try: w2.config(bg=color)
                    except: pass
            except: pass


if __name__ == "__main__":
    if DND_AVAILABLE:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()
    App(root)
    root.mainloop()
